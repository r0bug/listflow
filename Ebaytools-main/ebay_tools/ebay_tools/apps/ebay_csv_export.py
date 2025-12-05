#!/usr/bin/env python3
"""
ebay_csv_export.py - Export functionality for eBay listing data

This module provides functionality to export eBay listing data from JSON format
to both CSV and Excel formats suitable for eBay bulk upload.
"""

import os
import csv
import json
import sys
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from core.schema import EbayItemSchema, load_queue

# Excel support (optional, graceful fallback if not available)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
    WorkbookType = openpyxl.Workbook
    WorksheetType = openpyxl.worksheet.worksheet.Worksheet
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None
    # Define dummy types for type hints when openpyxl is not available
    class WorkbookType:
        pass
    class WorksheetType:
        pass


def load_json_queue(file_path: str) -> List[Dict[str, Any]]:
    """
    Load a queue of items from a JSON file.
    
    Args:
        file_path: Path to the JSON file
        
    Returns:
        List of item dictionaries
    """
    return load_queue(file_path)


def export_items_to_csv(items: List[Dict[str, Any]], 
                       output_file: str, 
                       default_values: Dict[str, str] = None,
                       description_dir: Optional[str] = None) -> Tuple[bool, str]:
    """
    Export items to CSV format suitable for eBay bulk upload.
    
    Args:
        items: List of item dictionaries to export
        output_file: Path to output CSV file
        default_values: Default values for CSV fields
        description_dir: Directory to save HTML description files
        
    Returns:
        Tuple of (success: bool, message: str)
    """
    try:
        if not items:
            return False, "No items to export"
        
        if default_values is None:
            default_values = {}
        
        # Convert all items to CSV format
        csv_rows = []
        for item in items:
            row = EbayItemSchema.to_csv_row(item, default_values)
            csv_rows.append(row)
        
        # Get all unique field names from all rows
        all_fields = set()
        for row in csv_rows:
            all_fields.update(row.keys())
        
        # Sort fields to put standard eBay fields first
        standard_fields = [
            "CustomLabel", "Title", "Category", "StartPrice", "Quantity", 
            "ConditionID", "ConditionDescription", "Format", "Duration"
        ]
        
        sorted_fields = []
        for field in standard_fields:
            if field in all_fields:
                sorted_fields.append(field)
                all_fields.remove(field)
        
        # Add remaining fields in alphabetical order
        sorted_fields.extend(sorted(all_fields))
        
        # Write CSV file
        with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=sorted_fields)
            writer.writeheader()
            writer.writerows(csv_rows)
        
        # Create HTML description files if requested
        created_descriptions = 0
        if description_dir:
            os.makedirs(description_dir, exist_ok=True)
            for item in items:
                if _create_html_description(item, description_dir):
                    created_descriptions += 1
        
        # Prepare success message
        message = f"Successfully exported {len(items)} items to {output_file}"
        if created_descriptions > 0:
            message += f"\nCreated {created_descriptions} HTML description files in {description_dir}"
        
        return True, message
        
    except Exception as e:
        return False, f"Error exporting to CSV: {str(e)}"


def export_items_to_excel(items: List[Dict[str, Any]], 
                         output_file: str, 
                         default_values: Dict[str, str] = None,
                         description_dir: Optional[str] = None) -> Tuple[bool, str]:
    """
    Export items to Excel format following eBay bulk upload specification.
    
    Args:
        items: List of item dictionaries to export
        output_file: Path to output Excel file
        default_values: Default values for fields
        description_dir: Directory to save HTML description files
        
    Returns:
        Tuple of (success: bool, message: str)
    """
    try:
        if not EXCEL_AVAILABLE:
            return False, "Excel export requires openpyxl package. Install with: pip install openpyxl"
        
        if not items:
            return False, "No items to export"
        
        if default_values is None:
            default_values = {}
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create required sheets
        _create_welcome_sheet(wb)
        _create_instructions_sheet(wb)
        main_sheet = _create_main_template_sheet(wb, items, default_values)
        
        # Save workbook
        wb.save(output_file)
        
        # Create HTML description files if requested
        created_descriptions = 0
        if description_dir:
            os.makedirs(description_dir, exist_ok=True)
            for item in items:
                if _create_html_description(item, description_dir):
                    created_descriptions += 1
        
        # Prepare success message
        message = f"Successfully exported {len(items)} items to Excel format: {output_file}"
        if created_descriptions > 0:
            message += f"\nCreated {created_descriptions} HTML description files in {description_dir}"
        
        return True, message
        
    except Exception as e:
        return False, f"Error exporting to Excel: {str(e)}"


def _create_welcome_sheet(wb: WorkbookType) -> None:
    """Create the WELCOME sheet for the eBay Excel template."""
    ws = wb.create_sheet("WELCOME")
    
    # Add welcome message
    ws['A1'] = "Welcome to eBay Tools Excel Export"
    ws['A1'].font = Font(size=16, bold=True)
    
    ws['A3'] = "This file has been generated by eBay Tools and follows the eBay bulk upload format specification."
    ws['A4'] = "The main data is in the 'eBay-prefill-listing-template' sheet."
    ws['A5'] = "Please review all data before uploading to eBay."
    
    ws['A7'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A8'] = "Generated by: eBay Tools"


def _create_instructions_sheet(wb: WorkbookType) -> None:
    """Create the GENERAL INSTRUCTIONS sheet for the eBay Excel template."""
    ws = wb.create_sheet("GENERAL INSTRUCTIONS")
    
    instructions = [
        "GENERAL INSTRUCTIONS FOR EBAY BULK UPLOAD",
        "",
        "1. Review all data in the 'eBay-prefill-listing-template' sheet",
        "2. Verify image URLs are accessible and properly formatted",
        "3. Check that titles are under 80 characters",
        "4. Ensure all required fields are completed",
        "5. Save the file in Excel (.xlsx) format",
        "6. Upload via eBay Seller Hub > Reports page",
        "",
        "SUPPORTED INPUT SETS:",
        "- Set A: Photo URL + Title (for photo-based listings)",
        "- Set B: Category + Aspects (for specification-based listings)",
        "- Set C: Product ID + Type (for ASIN/WID imports)",
        "- Set D: Item URL (for existing listing imports)",
        "",
        "NOTES:",
        "- At least one complete input set must be provided per item",
        "- Image URLs must start with https:// and end with valid extensions",
        "- Multiple image URLs should be separated by pipe (|) character",
        "- Maximum 24 images per listing",
        "- Review eBay policies before uploading"
    ]
    
    for i, instruction in enumerate(instructions, 1):
        ws[f'A{i}'] = instruction
        if i == 1:  # Title
            ws[f'A{i}'].font = Font(size=14, bold=True)


def _create_main_template_sheet(wb: WorkbookType, 
                               items: List[Dict[str, Any]], 
                               default_values: Dict[str, str]) -> WorksheetType:
    """Create the main eBay-prefill-listing-template sheet with data."""
    ws = wb.create_sheet("eBay-prefill-listing-template")
    
    # Add required header rows
    ws['A1'] = "#INFO | Version=1.0.0 | | Template=eBay-taxonomy-mapping-template_US"
    ws['A2'] = "#INFO | Set A | | Set B"
    
    # Define eBay column headers (per specification)
    headers = [
        "Custom Label (SKU)",      # A - SKU/identifier
        "Item Photo URL",          # B - Image URLs
        "Title",                   # C - Listing title
        "Category",                # D - Category or Product ID
        "Aspects",                 # E - Aspects or Product ID Type  
        "Item URL"                 # F - E-commerce URLs
    ]
    
    # Add headers in row 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Convert items to eBay format and add data
    for row_idx, item in enumerate(items, 4):  # Start from row 4
        # Map item data to eBay columns
        ebay_row = _convert_item_to_ebay_format(item, default_values)
        
        # Add data to columns
        ws.cell(row=row_idx, column=1, value=ebay_row.get("custom_label", ""))  # Custom Label (SKU)
        ws.cell(row=row_idx, column=2, value=ebay_row.get("photo_urls", ""))   # Item Photo URL
        ws.cell(row=row_idx, column=3, value=ebay_row.get("title", ""))        # Title
        ws.cell(row=row_idx, column=4, value=ebay_row.get("category", ""))     # Category
        ws.cell(row=row_idx, column=5, value=ebay_row.get("aspects", ""))      # Aspects
        ws.cell(row=row_idx, column=6, value=ebay_row.get("item_url", ""))     # Item URL
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return ws


def _convert_item_to_ebay_format(item: Dict[str, Any], 
                                default_values: Dict[str, str]) -> Dict[str, str]:
    """
    Convert an eBay Tools item to eBay bulk upload format.
    
    Args:
        item: Item dictionary from eBay Tools
        default_values: Default values to use
        
    Returns:
        Dictionary with eBay format fields
    """
    result = {}
    
    # Custom Label (SKU) - Column A
    result["custom_label"] = item.get("sku", "")
    
    # Item Photo URL - Column B  
    photo_urls = []
    if "photos" in item and isinstance(item["photos"], list):
        for photo in item["photos"]:
            if isinstance(photo, dict) and "path" in photo:
                # Convert local paths to URLs if needed
                path = photo["path"]
                if path.startswith("http"):
                    photo_urls.append(path)
                else:
                    # For local files, you might want to upload them first
                    # For now, we'll skip local files
                    pass
    
    result["photo_urls"] = "|".join(photo_urls) if photo_urls else ""
    
    # Title - Column C
    title = item.get("title", item.get("temp_title", ""))
    if title and len(title) > 80:
        title = title[:80]  # eBay limit
    result["title"] = title
    
    # Category - Column D (can be category name or product ID)
    result["category"] = item.get("category", "")
    
    # Aspects - Column E (item specifics in name=value|name2=value2 format)
    aspects = []
    item_specifics = EbayItemSchema.extract_item_specifics(item)
    
    for name, value in item_specifics.items():
        if name and value:
            # Clean the name and value for eBay format
            clean_name = str(name).replace("=", "").replace("|", "")
            clean_value = str(value).replace("=", "").replace("|", "")
            aspects.append(f"{clean_name}={clean_value}")
    
    # Add basic listing details as aspects
    if item.get("condition"):
        condition_name = EbayItemSchema.get_display_condition(item)
        aspects.append(f"Condition={condition_name}")
    
    result["aspects"] = "|".join(aspects) if aspects else ""
    
    # Item URL - Column F (for imports from other platforms)
    result["item_url"] = item.get("source_url", "")
    
    return result


def _create_html_description(item: Dict[str, Any], description_dir: str) -> bool:
    """
    Create an HTML description file for an item.
    
    Args:
        item: Item dictionary
        description_dir: Directory to save description files
        
    Returns:
        True if description was created, False otherwise
    """
    try:
        # Generate description content
        description = _generate_item_description(item)
        if not description:
            return False
        
        # Create filename based on SKU or ID
        sku = item.get("sku", item.get("id", "unknown"))
        filename = f"{sku}_description.html"
        filepath = os.path.join(description_dir, filename)
        
        # Write HTML file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(description)
        
        return True
        
    except Exception:
        return False


def _generate_item_description(item: Dict[str, Any]) -> str:
    """
    Generate HTML description for an item.
    
    Args:
        item: Item dictionary
        
    Returns:
        HTML description string
    """
    # Check if we have processed description from API results
    api_results = item.get("api_results", [])
    if api_results and isinstance(api_results, list):
        for result in api_results:
            if isinstance(result, dict) and "final_description" in result:
                return result["final_description"]
    
    # Fallback: generate basic description
    title = item.get("title", item.get("temp_title", ""))
    condition = EbayItemSchema.get_display_condition(item)
    item_specifics = EbayItemSchema.extract_item_specifics(item)
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; }}
        .specs {{ margin: 10px 0; }}
        .spec-name {{ font-weight: bold; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <p><strong>Condition:</strong> {condition}</p>
"""
    
    if item_specifics:
        html += "<h2>Item Specifics</h2>\n<div class='specs'>\n"
        for name, value in item_specifics.items():
            html += f"<div><span class='spec-name'>{name}:</span> {value}</div>\n"
        html += "</div>\n"
    
    html += "</body>\n</html>"
    
    return html